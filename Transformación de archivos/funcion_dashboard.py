import openpyxl
import pandas as pd
import openpyxl
import os
from collections import defaultdict

def DFaños (año,workbook):
  
  sheet = workbook[año]
  data = []
  headers = []
  columns_to_exclude = ['A']  
  start_row = 4

  column_counts = defaultdict(int)

  for i, row in enumerate(sheet.iter_rows(min_row=start_row)):
      row_data = []
      for j, cell in enumerate(row):
          if openpyxl.utils.get_column_letter(j+1) not in columns_to_exclude:
              if i == 0: 
                  column_name = cell.value
                  column_counts[column_name] += 1
                  if column_counts[column_name] > 1:
                      column_name = f"{column_name}_{column_counts[column_name]}"
                  headers.append(column_name)
              else:
                  row_data.append(cell.value)
      if row_data: 
          data.append(row_data)

  
  datos_año = pd.DataFrame(data, columns=headers)
  datos_año_copia = datos_año.copy()
  datos_año_copia = datos_año_copia.drop(['IR CUATRIMESTRE 1', 'IR CUATRIMESTRE 2', 'IR CUATRIMESTRE 3'], axis=1)

  df_reorganizado = pd.DataFrame(columns=['DESCRIPCIÓN DEL ARTICULO', 'MES', 'VALOR'])
  for index, row in datos_año_copia.iterrows():
      for mes in datos_año_copia.columns[1:]:
          df_reorganizado = df_reorganizado.append({
              'DESCRIPCIÓN DEL ARTICULO': row['DESCRIPCIÓN DEL ARTICULO'],
              'MES': mes,
              'VALOR': row[mes]
          }, ignore_index=True)
  def clasificar_tipo_operacion(mes):
      if '_2' in mes:
          return 'COMPRAS'
      elif '_3' in mes:
          return 'INVENTARIOS'
      else:
          return 'VENTAS'

  df_reorganizado['TIPO DE OPERACIÓN'] = df_reorganizado['MES'].apply(clasificar_tipo_operacion)
  df_reorganizado
  
  df_reorganizado['MES'] = df_reorganizado['MES'].str.replace(r"_[23]", "", regex=True)
  df_reorganizado['AÑO'] = año
  df_reorganizado = df_reorganizado.copy()
  return df_reorganizado
 


def procesarArhcivo(ruta,progress_queue):
    dfs = []
    workbook = openpyxl.load_workbook(ruta)
    sheet_names = workbook.sheetnames
    años = [sheet for sheet in sheet_names if sheet.isdigit() and int(sheet) >= 2020 and int(sheet) <= 2023]

    for año in años:
        df_año = DFaños(año,workbook)
        dfs.append(df_año)
    df_concatenado = pd.concat(dfs, ignore_index=True)
    df_concatenado.fillna(0, inplace=True)
    ruta_Archivo= os.path.dirname(ruta)+"/Dashboard.xlsx"
    df_concatenado.to_excel(ruta_Archivo, index=False)
    progress_queue.put("La extracción finalizó")
